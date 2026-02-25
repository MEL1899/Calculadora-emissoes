from flask import Flask, render_template, request, send_file
import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from carbon_calculator import (
    calcular_emissao, 
    calcular_fator_idade, 
    calcular_intensidade
)

# Tenta importar configuração do Google Maps API
# Em produção, usa variável de ambiente; em desenvolvimento, usa config.py
import os
try:
    from config import GOOGLE_MAPS_API_KEY
    # Se estiver vazio, tenta variável de ambiente
    if GOOGLE_MAPS_API_KEY == 'YOUR_API_KEY':
        GOOGLE_MAPS_API_KEY = os.getenv('GOOGLE_MAPS_API_KEY', 'YOUR_API_KEY')
except ImportError:
    # Se não houver config.py, usa variável de ambiente
    GOOGLE_MAPS_API_KEY = os.getenv('GOOGLE_MAPS_API_KEY', 'YOUR_API_KEY')

app = Flask(__name__)

# Estilos para o Excel
_borda_fina = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
_cabecalho_fill = PatternFill(start_color='2d5a3d', end_color='2d5a3d', fill_type='solid')
_cabecalho_font = Font(bold=True, color='FFFFFF', size=11)


def _formatar_tabela_resumo(ws):
    """Formata a planilha Resumo: largura das colunas, cabeçalho e bordas."""
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 28
    # Cabeçalho (linha 3)
    for col in ['A', 'B']:
        cell = ws[f'{col}3']
        cell.font = _cabecalho_font
        cell.fill = _cabecalho_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = _borda_fina
    # Dados (linhas 4 até max_row)
    for row in range(4, ws.max_row + 1):
        for col in ['A', 'B']:
            cell = ws[f'{col}{row}']
            cell.border = _borda_fina
            cell.alignment = Alignment(vertical='center', wrap_text=True)


def _formatar_sheet_tecnicos(ws):
    """Formata a planilha Dados técnicos: cabeçalho e bordas."""
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 18
    for cell in ws[1]:
        cell.font = _cabecalho_font
        cell.fill = _cabecalho_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _borda_fina
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = _borda_fina
            ws.cell(row=row, column=col).alignment = Alignment(vertical='center')


def _obter_consumo_medio_km_por_litro(tipo_veiculo: str, tipo_combustivel: str) -> float:
    """
    Retorna um consumo médio aproximado (km/L) por tipo de veículo e combustível.
    Valores apenas indicativos para cálculo estimado.
    """
    medias = {
        'carro': {
            'Gasolina': 12.0,
            'Etanol': 8.0,
            'Diesel S10': 14.0,
        },
        'caminhao': {
            'Diesel S10': 2.5,
            'Gasolina': 2.0,
            'Etanol': 1.8,
        }
    }
    tipo = medias.get(tipo_veiculo.lower(), {})
    valor = tipo.get(tipo_combustivel, 5.0)
    return float(valor) if valor > 0 else 5.0


@app.context_processor
def inject_google_maps_key():
    """Injeta a API key do Google Maps em todos os templates"""
    return dict(google_maps_api_key=GOOGLE_MAPS_API_KEY)

@app.route('/')
def home():
    return render_template('index.html', ano_atual=datetime.now().year)

@app.route('/calcular', methods=['POST'])
def calcular():
    try:
        # 1. Captura todos os dados do formulário
        email = request.form['email']
        tipo_veiculo = request.form.get('tipo_veiculo', 'carro')  # Default para carro
        tipo_combustivel = request.form['tipo_combustivel']
        modo_calculo = request.form.get('modo_calculo', 'preciso')
        ano_fabricacao = int(request.form['ano_fabricacao'])
        km_rodado = float(request.form.get('km_rodado', 0) or 0)
        # Campo de carga é opcional para carros (será 0 se não fornecido)
        carga_ton = float(request.form.get('carga_ton', 0) or 0)

        # Define litros conforme o modo de cálculo
        litros_estimado = False
        if modo_calculo == 'preciso':
            litros = float(request.form['litros'])
        else:
            # Estimativa de litros com base em km rodado e consumo médio
            consumo_medio = _obter_consumo_medio_km_por_litro(tipo_veiculo, tipo_combustivel)
            if km_rodado <= 0 or consumo_medio <= 0:
                litros = 0.0
            else:
                litros = km_rodado / consumo_medio
            litros_estimado = True
        
        # Dados de localização (opcionais)
        endereco_origem = request.form.get('endereco_origem', '')
        endereco_destino = request.form.get('endereco_destino', '')
        lat_origem = request.form.get('lat_origem', '')
        lng_origem = request.form.get('lng_origem', '')
        lat_destino = request.form.get('lat_destino', '')
        lng_destino = request.form.get('lng_destino', '')
        
        # 2. Cria um DataFrame temporário com os dados do formulário
        # Isso permite usar as funções avançadas que esperam DataFrame
        dados_viagem = pd.DataFrame({
            'ID_Viagem': ['WEB_001'],
            'Data': [datetime.now().strftime('%Y-%m-%d')],
            'Frota_ID': ['WEB'],
            'Combustivel_L': [litros],
            'Tipo_Combustivel': [tipo_combustivel],
            'KM_Rodado': [km_rodado],
            'Carga_Ton': [carga_ton],
            'Numero_Eixos': [0],  # Não usado no cálculo, mas necessário para compatibilidade
            'Ano_Fabricacao': [ano_fabricacao]
        })
        
        # 3. Processa usando todas as funções avançadas
        
        # 3.1. Calcula emissão base
        dados_com_emissao = calcular_emissao(dados_viagem)
        
        # 3.2. Calcula fator de idade e emissão final
        dados_com_idade = calcular_fator_idade(dados_com_emissao)
        
        # 3.3. Calcula intensidades e eficiência
        dados_final = calcular_intensidade(dados_com_idade)
        
        # 4. Extrai os resultados para exibição
        resultado = dados_final.iloc[0]  # Primeira (e única) linha
        ano_atual = datetime.now().year
        idade_veiculo = ano_atual - ano_fabricacao
        tipo_veiculo_label = 'Caminhão' if tipo_veiculo == 'caminhao' else 'Carro'
        modo_calculo_label = 'Modo preciso' if modo_calculo == 'preciso' else 'Modo estimado'
        modo_calculo_selo = 'Alta precisão' if modo_calculo == 'preciso' else 'Resultado estimado (consumo médio)'
        litros_texto_origem = 'informados pelo usuário' if not litros_estimado else 'estimados a partir dos km rodados'

        # Comparação com média de intensidade do setor (valor fixo simulado)
        media_setor_intensidade_km = 0.0008  # tCO2e/km (valor de referência fixo)
        intensidade_km_valor = float(resultado['Intensidade_tCO2e_por_KM'])
        if km_rodado > 0 and intensidade_km_valor > 0 and media_setor_intensidade_km > 0:
            diff_percent = (intensidade_km_valor - media_setor_intensidade_km) / media_setor_intensidade_km * 100
            if diff_percent > 5:
                situacao_setor = 'acima'
            elif diff_percent < -5:
                situacao_setor = 'abaixo'
            else:
                situacao_setor = 'na faixa da'
            diff_percent_formatado = f"{abs(diff_percent):.1f}"
        else:
            situacao_setor = 'indisponível'
            diff_percent_formatado = None
        
        # 5. Gera relatório Excel individual (Resumo em português + Dados técnicos)
        nome_arquivo = f"Relatorio_Carbono_Individual_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho_arquivo = os.path.join(os.getcwd(), nome_arquivo)
        
        # Planilha Resumo: relatório legível em português
        resumo = pd.DataFrame({
            'Item': [
                'Tipo de veículo', 'Combustível', 'Litros consumidos (L)', 'Ano de fabricação',
                'Idade do veículo (anos)', 'Quilômetros rodados (km)', 'Emissão base (tCO2e)',
                'Fator de idade', 'Emissão final (tCO2e)', 'Intensidade por km (tCO2e/km)',
                'Eficiência (km/L)', 'E-mail', 'Data do relatório'
            ],
            'Valor': [
                tipo_veiculo_label, tipo_combustivel, f'{litros:.2f}', str(ano_fabricacao),
                str(idade_veiculo), f'{km_rodado:.2f}', f"{resultado['emissao_base']:.4f}",
                f"{resultado['Fator_idade']:.4f}", f"{resultado['emissao_final']:.4f}",
                f"{resultado['Intensidade_tCO2e_por_KM']:.6f}", f"{resultado['Eficiencia_KM_por_L']:.2f}",
                email, datetime.now().strftime('%d/%m/%Y %H:%M')
            ]
        })
        if tipo_veiculo == 'caminhao':
            idx_efic = resumo[resumo['Item'] == 'Eficiência (km/L)'].index[0]
            nova_linha = pd.DataFrame({'Item': ['Carga transportada (toneladas)', 'Intensidade por tonelada (tCO2e/ton)'], 'Valor': [f'{carga_ton:.2f}', f"{resultado['Intensidade_tCO2e_por_Ton']:.6f}"]})
            resumo = pd.concat([resumo.iloc[:idx_efic], nova_linha, resumo.iloc[idx_efic:]]).reset_index(drop=True)
        
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            # Escreve Resumo: título na linha 1, tabela a partir da linha 3
            resumo.to_excel(writer, sheet_name='Resumo', index=False, startrow=2)
            ws_resumo = writer.sheets['Resumo']
            ws_resumo['A1'] = 'Relatório de Emissões de GEE'
            ws_resumo['A1'].font = Font(bold=True, size=14)
            ws_resumo.merge_cells('A1:B1')
            ws_resumo['A2'] = f'Emitido em {datetime.now().strftime("%d/%m/%Y às %H:%M")} — Carbon Log'
            ws_resumo['A2'].font = Font(size=10, color='666666')
            ws_resumo.merge_cells('A2:B2')
            _formatar_tabela_resumo(ws_resumo)
            # Escreve Dados técnicos
            dados_final.to_excel(writer, sheet_name='Dados técnicos', index=False)
            _formatar_sheet_tecnicos(writer.sheets['Dados técnicos'])
        
        # 6. Prepara dados para o template
        data_relatorio = datetime.now().strftime('%d/%m/%Y às %H:%M')
        dados_template = {
            'data_relatorio': data_relatorio,
            'modo_calculo': modo_calculo,
            'modo_calculo_label': modo_calculo_label,
            'modo_calculo_selo': modo_calculo_selo,
            'litros_estimado': litros_estimado,
            'litros_texto_origem': litros_texto_origem,
            'email': email,
            'tipo_veiculo': tipo_veiculo_label,
            'tipo_veiculo_raw': tipo_veiculo,
            'tipo_combustivel': tipo_combustivel,
            'litros': f"{litros:.2f}",
            'ano_fabricacao': ano_fabricacao,
            'idade_veiculo': idade_veiculo,
            'km_rodado': f"{km_rodado:.2f}",
            'carga_ton': f"{carga_ton:.2f}",
            'endereco_origem': endereco_origem if endereco_origem else 'Não informado',
            'endereco_destino': endereco_destino if endereco_destino else 'Não informado',
            'emissao_base': f"{resultado['emissao_base']:.4f}",
            'fator_idade': f"{resultado['Fator_idade']:.4f}",
            'emissao_final': f"{resultado['emissao_final']:.4f}",
            'intensidade_ton': f"{resultado['Intensidade_tCO2e_por_Ton']:.6f}",
            'intensidade_km': f"{resultado['Intensidade_tCO2e_por_KM']:.6f}",
            'eficiencia': f"{resultado['Eficiencia_KM_por_L']:.2f}",
            'media_setor_intensidade_km': f"{media_setor_intensidade_km:.6f}",
            'comparacao_setor_percentual': diff_percent_formatado,
            'comparacao_setor_situacao': situacao_setor,
            'arquivo_relatorio': f"/download/{nome_arquivo}"
        }
        
        # 7. Renderiza a página de resultado
        return render_template('resultado.html', **dados_template)
        
    except Exception as e:
        # Tratamento de erro
        error_message = f"Erro ao processar o cálculo: {str(e)}"
        return f"<h1>Erro</h1><p>{error_message}</p><a href='/'>Voltar</a>", 400

@app.route('/download/<filename>')
def download_file(filename):
    """Rota para download do relatório Excel"""
    try:
        caminho_arquivo = os.path.join(os.getcwd(), filename)
        if os.path.exists(caminho_arquivo):
            return send_file(caminho_arquivo, as_attachment=True)
        else:
            return "Arquivo não encontrado", 404
    except Exception as e:
        return f"Erro ao fazer download: {str(e)}", 500

if __name__ == '__main__':
    app.run(debug=True)